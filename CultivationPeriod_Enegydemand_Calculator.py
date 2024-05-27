'''
This script reads the hourly greenhouse energy demand for heating and artificial lighting from an Excel file 
(Greenhouse energy_demand 2021.xlsx).

It first calculates the monthly energy demand and then allocates the heating and lighting energy demand 
to the harvesting periods based on the greenhouse crop growing cycles (in other words, it allocates the Energy demand 
for the pre-harvest period to the greenhouse harvesting period with different shares).

Finally, it writes the allocated heating and artificial lighting energy demand values for 
the greenhouse harvesting period to an output Excel file (filename).
'''


# Import necessary libraries
import pandas as PD
import matplotlib as PLT
import math as MS
import numpy as NP
from openpyxl import Workbook
from openpyxl import load_workbook  

# Introduce Countries & Crops
print("=============================================================")
print("Countries: Netherlands= NL, Belgium= BE, France=FR, Switzerland= CH")
print("Crops: Tomato, Cucumber, Lettuce, Bell pepper, and Strawberry")
print("=============================================================")


COUNTRY=['NL','BE','FR','CH']
CROPS={
   'NL':['Tomato', 'Cucumber', 'Lettuce', 'Bell pepper', 'Strawberry'],
   'BE':['Tomato', 'Lettuce', 'Strawberry'],
   'FR':['Tomato', 'Lettuce',],
   'CH':['Tomato', 'Cucumber', 'Lettuce', 'Bell pepper', 'Strawberry'] 
}

year='2021'

# Output is "Carbon footprint.xlsx" / include: Monthly carbon footprint according to the cultivation period of crops.
# The Excell file (Carbon footprint.xlsx) is important.
filename = "C:/Import VS Local/Results/Carbon footprint.xlsx"

for i in COUNTRY:
    for j in CROPS[i]:
        
        Country=i
        Crop=j
        
        # Hourly energy demand (Heat & Light) is read from the Excel file (Greenhouse Energy Demand_2021.xlsx).
        Energy_demand = 'C:/Import VS Local/Results/Greenhouse Energy Demand_2021.xlsx'  

        Heat=NP.array(PD.read_excel(Energy_demand, sheet_name=i+'_'+year+'_'+j, usecols='F'))
        Light=NP.array(PD.read_excel(Energy_demand, sheet_name=i+'_'+year+'_'+j, usecols='E'))
        
           

        # Monthly Heat demand(MJ//ha)
        Heat_1=sum(Heat[0:(31*24)-1])*3.6           # Jan
        Heat_2=sum(Heat[(31*24):(59*24)-1])*3.6     # Feb
        Heat_3=sum(Heat[(59*24):(90*24)-1])*3.6     # Mar
        Heat_4=sum(Heat[(90*24):(120*24)-1])*3.6    # Apr
        Heat_5=sum(Heat[(120*24):(151*24)-1])*3.6   # May
        Heat_6=sum(Heat[(151*24):(181*24)-1])*3.6   # Jun
        Heat_7=sum(Heat[(181*24):(212*24)-1])*3.6   # Jul
        Heat_8=sum(Heat[(212*24):(243*24)-1])*3.6   # Aug
        Heat_9=sum(Heat[(243*24):(273*24)-1])*3.6   # Sep
        Heat_10=sum(Heat[(273*24):(304*24)-1])*3.6  # Oct
        Heat_11=sum(Heat[(304*24):(334*24)-1])*3.6  # Nov
        Heat_12=sum(Heat[(334*24):(365*24)-1])*3.6  # Dec

        
        # Monthly Light demand (kWh/ha)
        Light_1=sum(Light[0:(31*24)-1])             # Jan
        Light_2=sum(Light[(31*24):(59*24)-1])       # Feb
        Light_3=sum(Light[(59*24):(90*24)-1])       # Mar
        Light_4=sum(Light[(90*24):(120*24)-1])      # Apr
        Light_5=sum(Light[(120*24):(151*24)-1])     # May
        Light_6=sum(Light[(151*24):(181*24)-1])     # Jun
        Light_7=sum(Light[(181*24):(212*24)-1])     # Jul
        Light_8=sum(Light[(212*24):(243*24)-1])     # Aug
        Light_9=sum(Light[(243*24):(273*24)-1])     # Sep
        Light_10=sum(Light[(273*24):(304*24)-1])    # Oct
        Light_11=sum(Light[(304*24):(334*24)-1])    # Nov
        Light_12=sum(Light[(334*24):(365*24)-1])    # Dec


        def write_data_to_excel(filename, sheetname, cell_ranges, data):
            try:
                # Load the workbook if it exists
                wb = load_workbook(filename)
            except FileNotFoundError:
                # Create a new workbook if the file does not exist
                wb = Workbook()
            
            # Select the desired sheet or create it if it does not exist
            if sheetname in wb.sheetnames:
                ws = wb[sheetname]
            else:
                ws = wb.create_sheet(title=sheetname)
            
            # Writing numbers and strings to specified cells
            for cell_range, item in zip(cell_ranges, data):
                ws[cell_range] = item
                        
            # Save the workbook
            wb.save(filename)


        
        #________________________________________________________________#
        
        ###### Cultivation periods related to different crops ######

        # Tomato (Cultivation period 1) (Cutivation period = 11 month and harvesting period = 9 month)
        if Crop=="Tomato" and Country in ["NL", "BE", "FR", "CH"]:
            # Cultivation period 1 (Fram Jan. to Nov.)
            
            # Heating demand pre-harvesting and harvesting month
            H1_Pre_1=Heat_1                           # Jan  (Pre-harvesting month)
            H1_Pre_2=Heat_2                           # Feb  (Pre-harvesting month)
            H1_harv_1=(H1_Pre_1+H1_Pre_2/2+Heat_3/3)  # Mar = Jan + 1/2 Feb + 1/3 Mar
            H1_harv_2=(H1_Pre_2/2+Heat_3/3+Heat_4/3)  # Apr = 1/2 Feb + 1/3 Mar + 1/3 Apr
            H1_harv_3=(Heat_3/3+Heat_4/3+Heat_5/3)    # May = 1/3 Mar + 1/3 Apr + 1/3 May
            H1_harv_4=(Heat_4/3+Heat_5/3+Heat_6/3)    # Jun = 1/3 Apr + 1/3 May + 1/3 Jun
            H1_harv_5=(Heat_5/3+Heat_6/3+Heat_7/3)    # Jul = 1/3 May + 1/3 Jun + 1/3 Jul
            H1_harv_6=(Heat_6/3+Heat_7/3+Heat_8/3)    # Aug = 1/3 Jun + 1/3 Jul + 1/3 Aug
            H1_harv_7=(Heat_7/3+Heat_8/3+Heat_9/3)    # Sep = 1/3 Jul + 1/3 Aug + 1/3 Sep
            H1_harv_8=(Heat_8/3+Heat_9/3+Heat_10/2)   # Oct = 1/3 Aug + 1/3 Sep + 1/2 Oct
            H1_harv_9=(Heat_9/3+Heat_10/2+Heat_11)    # Nov = 1/3 Sep + 1/2 Oct + Nov
            
            CP1_heat_tomato=[H1_Pre_1, H1_Pre_2, H1_harv_1, H1_harv_2, H1_harv_3, H1_harv_4, 
                             H1_harv_5, H1_harv_6, H1_harv_7, H1_harv_8, H1_harv_9]
        
            HEAT = []
            for sublist in CP1_heat_tomato[2::]:
                HEAT.append(sublist[0])
            CP1_HEAT_TOMATO = [round(number, 2) for number in HEAT]

            # lighting demand pre-harvesting and harvesting month
            L1_Pre_1=Light_1                              # Jan (Pre-harvesting month)
            L1_Pre_2=Light_2                              # Feb (Pre-harvesting month)
            L1_harv_1=(L1_Pre_1+L1_Pre_2/2+Light_3/3)     # Mar = Jan + 1/2 Feb + 1/3 Mar
            L1_harv_2=(L1_Pre_2/2+Light_3/3+Light_4/3)    # Apr = 1/2 Feb + 1/3 Mar + 1/3 Apr
            L1_harv_3=(Light_3/3+Light_4/3+Light_5/3)     # May = 1/3 Mar + 1/3 Apr + 1/3 May
            L1_harv_4=(Light_4/3+Light_5/3+Light_6/3)     # Jun = 1/3 Apr + 1/3 May + 1/3 Jun
            L1_harv_5=(Light_5/3+Light_6/3+Light_7/3)     # Jul = 1/3 May + 1/3 Jun + 1/3 Jul
            L1_harv_6=(Light_6/3+Light_7/3+Light_8/3)     # Aug = 1/3 Jun + 1/3 Jul + 1/3 Aug
            L1_harv_7=(Light_7/3+Light_8/3+Light_9/3)     # Sep = 1/3 Jul + 1/3 Aug + 1/3 Sep
            L1_harv_8=(Light_8/3+Light_9/3+Light_10/2)    # Oct = 1/3 Aug + 1/3 Sep + 1/2 Oct
            L1_harv_9=(Light_9/3+Light_10/2+Light_11)     # Nov = 1/3 Sep + 1/2 Oct + Nov
            
            CP1_light_tomato=[L1_Pre_1, L1_Pre_2, L1_harv_1, L1_harv_2, L1_harv_3, L1_harv_4,
                              L1_harv_5, L1_harv_6, L1_harv_7, L1_harv_8, L1_harv_9]
            
            LIGHT = []
            for sublist in CP1_light_tomato[2::]:
                LIGHT.append(sublist[0])
            CP1_LIGHT_TOMATO = [round(number, 2) for number in LIGHT]
            
            
            # Writing to the desired cells in Excel
            if Country =="NL":
                sheetname = "NL_Tomato"
            elif Country =="BE":     
                sheetname = "BE_Tomato"
            elif Country =="FR":     
                sheetname = "FR_Tomato"
            elif Country =="CH":     
                sheetname = "CH_Tomato"
            cell_ranges = ["B5","C5","C6","C7","D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5", "N5", "O5"]
            Month = ["Cultivation period1","Energy demand","Greenhouse heating demand(MJ/ha)","Greenhouse lighting demand(kWh/ha)","Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            write_data_to_excel(filename, sheetname, cell_ranges, Month) 
            cell_ranges = ["F6", "G6", "H6", "I6", "J6", "K6", "L6", "M6", "N6"]
            CP1_Heat_Tomato= CP1_HEAT_TOMATO  
            write_data_to_excel(filename, sheetname, cell_ranges,CP1_Heat_Tomato )
            cell_ranges = ["F7", "G7", "H7", "I7", "J7", "K7", "L7", "M7", "N7"]
            CP1_Light_Tomato= CP1_LIGHT_TOMATO  
            write_data_to_excel(filename, sheetname, cell_ranges,CP1_Light_Tomato)   
                
            print('done', i+'_'+j)
        
            #################################################
            
        # Tomato (Cultivation period 2) (Cutivation period = 11 month and harvesting period = 9 month)
            
        if Crop=="Tomato" and Country in ["NL", "BE", "CH"]:
            # Cultivation period 2 (from Oct. to Aug.)
            
            # Heating demand pre-harvesting and harvesting month (Cultivation period 2)
            H2_Pre_1=Heat_10                              # Oct (Pre-harvesting month)
            H2_Pre_2=Heat_11                              # Nov (Pre-harvesting month)
            H2_harv_1=(H2_Pre_1+H2_Pre_2/2+Heat_12/3)     # Dec = Oct + 1/2 Nov + 1/3 Dec
            H2_harv_2=(H2_Pre_2/2+Heat_12/3+Heat_1/3)     # Jan = 1/2 Nov + 1/3 Dec + 1/3 Jan
            H2_harv_3=(Heat_12/3+Heat_1/3+Heat_2/3)       # Feb = 1/3 Dec + 1/3 Jan + 1/3 Feb
            H2_harv_4=(Heat_1/3+Heat_2/3+Heat_3/3)        # Mar = 1/3 Jan + 1/3 Feb + 1/3 Mar
            H2_harv_5=(Heat_2/3+Heat_3/3+Heat_4/3)        # Apr = 1/3 Feb + 1/3 Mar + 1/3 Apr
            H2_harv_6=(Heat_3/3+Heat_4/3+Heat_5/3)        # May = 1/3 Mar + 1/3 Apr + 1/3 May
            H2_harv_7=(Heat_4/3+Heat_5/3+Heat_6/3)        # Jun = 1/3 Apr + 1/3 May + 1/3 Jun
            H2_harv_8=(Heat_5/3+Heat_6/3+Heat_7/2)        # Jul = 1/3 May + 1/3 Jun + 1/2 Jul
            H2_harv_9=(Heat_6/3+Heat_7/2+Heat_8)          # Aug = 1/3 Jun + 1/2 Jul + Aug
            
            CP2_heat_tomato=[H2_Pre_1, H2_Pre_2, H2_harv_1, H2_harv_2, H2_harv_3, H2_harv_4,
                            H2_harv_5, H2_harv_6, H2_harv_7, H2_harv_8, H2_harv_9]
            
            HEAT = []
            for sublist in CP2_heat_tomato[2::]:
                HEAT.append(sublist[0])
            CP2_HEAT_TOMATO = [round(number, 2) for number in HEAT]
            
            
            # lighting demand pre-harvesting and harvesting month (Cultivation period 2)
            L2_Pre_1=Light_10                            # Oct (Pre-harvesting month)
            L2_Pre_2=Light_11                            # Nov (Pre-harvesting month)
            L2_harv_1=(L2_Pre_1+L2_Pre_2/2+Light_12/3)   # Dec = Oct + 1/2 Nov + 1/3 Dec
            L2_harv_2=(L1_Pre_2/2+Light_12/3+Light_1/3)  # Jan = 1/2 Nov + 1/3 Dec + 1/3 Jan
            L2_harv_3=(Light_12/3+Light_1/3+Light_2/3)   # Feb = 1/3 Dec + 1/3 Jan + 1/3 Feb
            L2_harv_4=(Light_1/3+Light_2/3+Light_3/3)    # Mar = 1/3 Jan + 1/3 Feb + 1/3 Mar
            L2_harv_5=(Light_2/3+Light_3/3+Light_4/3)    # Apr = 1/3 Feb + 1/3 Mar + 1/3 Apr
            L2_harv_6=(Light_3/3+Light_4/3+Light_5/3)    # May = 1/3 Mar + 1/3 Apr + 1/3 May
            L2_harv_7=(Light_4/3+Light_5/3+Light_6/3)    # Jun = 1/3 Apr + 1/3 May + 1/3 Jun
            L2_harv_8=(Light_5/3+Light_6/3+Light_7/2)    # Jul = 1/3 May + 1/3 Jun + 1/2 Jul
            L2_harv_9=(Light_6/3+Light_7/2+Light_8)      # Aug = 1/3 Jun + 1/2 Jul + Aug
            
            CP2_light_tomato=[L2_Pre_1, L2_Pre_2, L2_harv_1, L2_harv_2, L2_harv_3, L2_harv_4,
                            L2_harv_5, L2_harv_6, L2_harv_7, L2_harv_8, L2_harv_9]
            
            LIGHT = []
            for sublist in CP2_light_tomato[2::]:
                LIGHT.append(sublist[0])
            CP2_LIGHT_TOMATO = [round(number, 2) for number in LIGHT]

            # Writing to the desired cells in Excel
            if Country =="NL":
                sheetname = "NL_Tomato"
            elif Country =="BE":     
                sheetname = "BE_Tomato"
            elif Country =="CH":     
                sheetname = "CH_Tomato"
            cell_ranges = ["B11","C11","C12","C13","D11", "E11", "F11", "G11", "H11", "I11", "J11", "K11", "L11", "M11", "N11", "O11"]
            Month = ["Cultivation period2","Energy demand","Greenhouse heating demand(MJ/ha)","Greenhouse lighting demand(kWh/ha)","Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            write_data_to_excel(filename, sheetname, cell_ranges, Month) 
            cell_ranges = ["O12", "D12", "E12", "F12", "G12", "H12", "I12", "J12", "K12"]
            CP2_Heat_Tomato= CP2_HEAT_TOMATO  #[H1_harv_1, H1_harv_2, H1_harv_3, H1_harv_4, H1_harv_5, H1_harv_6, H1_harv_7, H1_harv_8, H1_harv_9]
            write_data_to_excel(filename, sheetname, cell_ranges,CP2_Heat_Tomato )
            cell_ranges = ["O13", "D13", "E13", "F13", "G13", "H13", "I13", "J13", "K13"]
            CP2_Light_Tomato= CP2_LIGHT_TOMATO  #[H1_harv_1, H1_harv_2, H1_harv_3, H1_harv_4, H1_harv_5, H1_harv_6, H1_harv_7, H1_harv_8, H1_harv_9]
            write_data_to_excel(filename, sheetname, cell_ranges,CP2_Light_Tomato)   
                
            print('done', i+'_'+j+'2') 



        ###############################################
        ###############################################
        ###############################################

        # Bell pepper (Cutivation period = 11 month and harvesting period = 9 month)
        
        if Crop=="Bell pepper" and Country in ["NL", "CH"]:    
            
            # Heating demand pre-harvesting and harvesting month ()
            H1_Pre_1=Heat_1                           # Jan  (Pre-harvesting month)
            H1_Pre_2=Heat_2                           # Feb  (Pre-harvesting month)
            H1_harv_1=(H1_Pre_1+H1_Pre_2/2+Heat_3/3)  # Mar = Jan + 1/2 Feb + 1/3 Mar
            H1_harv_2=(H1_Pre_2/2+Heat_3/3+Heat_4/3)  # Apr = 1/2 Feb + 1/3 Mar + 1/3 Apr
            H1_harv_3=(Heat_3/3+Heat_4/3+Heat_5/3)    # May = 1/3 Mar + 1/3 Apr + 1/3 May
            H1_harv_4=(Heat_4/3+Heat_5/3+Heat_6/3)    # Jun = 1/3 Apr + 1/3 May + 1/3 Jun
            H1_harv_5=(Heat_5/3+Heat_6/3+Heat_7/3)    # Jul = 1/3 May + 1/3 Jun + 1/3 Jul
            H1_harv_6=(Heat_6/3+Heat_7/3+Heat_8/3)    # Aug = 1/3 Jun + 1/3 Jul + 1/3 Aug
            H1_harv_7=(Heat_7/3+Heat_8/3+Heat_9/3)    # Sep = 1/3 Jul + 1/3 Aug + 1/3 Sep
            H1_harv_8=(Heat_8/3+Heat_9/3+Heat_10/2)   # Oct = 1/3 Aug + 1/3 Sep + 1/2 Oct
            H1_harv_9=(Heat_9/3+Heat_10/2+Heat_11)    # Nov = 1/3 Sep + 1/2 Oct + Nov
            
            CP_heat_bell=[H1_Pre_1, H1_Pre_2, H1_harv_1, H1_harv_2, H1_harv_3, H1_harv_4, 
                             H1_harv_5, H1_harv_6, H1_harv_7, H1_harv_8, H1_harv_9]
            
            
            HEAT = []
            for sublist in CP_heat_bell[2::]:
                HEAT.append(sublist[0])
            CP_HEAT_BELL = [round(number, 2) for number in HEAT]
            
            
            # lighting demand pre-harvesting and harvesting month
            L1_Pre_1=Light_1                              # Jan (Pre-harvesting month)
            L1_Pre_2=Light_2                              # Feb (Pre-harvesting month)
            L1_harv_1=(L1_Pre_1+L1_Pre_2/2+Light_3/3)     # Mar = Jan + 1/2 Feb + 1/3 Mar
            L1_harv_2=(L1_Pre_2/2+Light_3/3+Light_4/3)    # Apr = 1/2 Feb + 1/3 Mar + 1/3 Apr
            L1_harv_3=(Light_3/3+Light_4/3+Light_5/3)     # May = 1/3 Mar + 1/3 Apr + 1/3 May
            L1_harv_4=(Light_4/3+Light_5/3+Light_6/3)     # Jun = 1/3 Apr + 1/3 May + 1/3 Jun
            L1_harv_5=(Light_5/3+Light_6/3+Light_7/3)     # Jul = 1/3 May + 1/3 Jun + 1/3 Jul
            L1_harv_6=(Light_6/3+Light_7/3+Light_8/3)     # Aug = 1/3 Jun + 1/3 Jul + 1/3 Aug
            L1_harv_7=(Light_7/3+Light_8/3+Light_9/3)     # Sep = 1/3 Jul + 1/3 Aug + 1/3 Sep
            L1_harv_8=(Light_8/3+Light_9/3+Light_10/2)    # Oct = 1/3 Aug + 1/3 Sep + 1/2 Oct
            L1_harv_9=(Light_9/3+Light_10/2+Light_11)     # Nov = 1/3 Sep + 1/2 Oct + Nov
            
            CP_light_bell=[L1_Pre_1, L1_Pre_2, L1_harv_1, L1_harv_2, L1_harv_3, L1_harv_4,
                              L1_harv_5, L1_harv_6, L1_harv_7, L1_harv_8, L1_harv_9]


            LIGHT = []
            for sublist in CP_light_bell[2::]:
                LIGHT.append(sublist[0])
            CP_LIGHT_BELL= [round(number, 2) for number in LIGHT]

            # Writing to the desired cells in Excel
            if Country =="NL":
                sheetname = "NL_Bell pepper"
            elif Country =="CH":     
                sheetname = "CH_Bell pepper"
            cell_ranges = ["B5","C5","C6","C7","D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5", "N5", "O5"]
            Month = ["Cultivation period2","Energy demand","Greenhouse heating demand(MJ/ha)","Greenhouse lighting demand(kWh/ha)","Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            write_data_to_excel(filename, sheetname, cell_ranges, Month) 
            cell_ranges = ["F6", "G6", "H6", "I6", "J6", "K6", "L6", "M6", "N6"]
            CP_Heat_Bell= CP_HEAT_BELL  #[H1_harv_1, H1_harv_2, H1_harv_3, H1_harv_4, H1_harv_5, H1_harv_6, H1_harv_7, H1_harv_8, H1_harv_9]
            write_data_to_excel(filename, sheetname, cell_ranges,CP_Heat_Bell)
            cell_ranges = ["F7", "G7", "H7", "I7", "J7", "K7", "L7", "M7", "N7"]
            CP_Light_Bell= CP_LIGHT_BELL  #[H1_harv_1, H1_harv_2, H1_harv_3, H1_harv_4, H1_harv_5, H1_harv_6, H1_harv_7, H1_harv_8, H1_harv_9]
            write_data_to_excel(filename, sheetname, cell_ranges,CP_Light_Bell)   
                
            print('done', i+'_'+j) 


        
        ###########################################################
        ###########################################################
        ###########################################################
        
        # Lettuce (Cultivation period ~ 45-60 days (2 month))
        
        if Crop=="Lettuce" and Country in ["NL", "BE", "FR", "CH"]:
            # Heating demand pre-harvesting and harvesting month
            H_Pre_1=Heat_1         # Jan (Pre-harvesting month)
            H_harv_1=H_Pre_1+Heat_2     # Feb = Jan + Feb
            
            H_Pre_2=Heat_2         # Feb (Pre-harvesting month)
            H_harv_2=H_Pre_2+Heat_3     # Mar = Feb + Mar
            
            H_Pre_3=Heat_3         # Mar (Pre-harvesting month)
            H_harv_3=H_Pre_3+Heat_4     # Apr = Mar + Apr 
            
            H_Pre_4=Heat_4         # Apr (Pre-harvesting month)
            H_harv_4=H_Pre_4+Heat_5     # May = Apr + May
            
            H_Pre_5=Heat_5         # May (Pre-harvesting month)
            H_harv_5=H_Pre_5+Heat_6     # Jun = May + Jun
            
            H_Pre_6=Heat_6         # Jun (Pre-harvesting month)
            H_harv_6=H_Pre_6+Heat_7     # Jul = Jun + Jul
            
            H_Pre_7=Heat_7         # Jul (Pre-harvesting month)
            H_harv_7=H_Pre_7+Heat_8     # Aug = Jul + Aug
                
            H_Pre_8=Heat_8         # Aug (Pre-harvesting month)
            H_harv_8=H_Pre_8+Heat_9     # Sep = Aug + Sep
            
            H_Pre_9=Heat_9         # Sep (Pre-harvesting month)
            H_harv_9=H_Pre_9+Heat_10    # Oct = Sep + Oct
            
            H_Pre_10=Heat_10       # Oct (Pre-harvesting month)
            H_harv_10=H_Pre_10+Heat_11  # Nov = Oct + Nov
            
            H_Pre_11=Heat_11       # Nov (Pre-harvesting month)
            H_harv_11=H_Pre_11+Heat_12  # Dec = Nov + Dec
            #H_Pre_12=Heat_12
            #H_harv_12=H_Pre_12+Heat_1   # Jan.
            CP_heat_lettuce=[H_harv_1, H_harv_2, H_harv_3, H_harv_4, H_harv_5, H_harv_6, 
                            H_harv_7, H_harv_8, H_harv_9, H_harv_10, H_harv_11]
            
            HEAT = []
            for sublist in CP_heat_lettuce[:]:
                HEAT.append(sublist[0])
            CP_HEAT_LETTUCE = [round(number, 2) for number in HEAT]
            
            
            # lighting demand pre-harvesting and harvesting month
            L_Pre_1=Light_1         # Jan (Pre-harvesting month)
            L_harv_1=L_Pre_1+Light_2     # Feb = Jan + Feb
            
            L_Pre_2=Light_2         # Feb (Pre-harvesting month)
            L_harv_2=L_Pre_2+Light_3     # Mar = Feb + Mar
            
            L_Pre_3=Light_3         # Mar (Pre-harvesting month)
            L_harv_3=L_Pre_3+Light_4     # Apr = Mar + Apr
            
            L_Pre_4=Light_4         # Apr (Pre-harvesting month)
            L_harv_4=L_Pre_4+Light_5     # May = Apr + May
             
            L_Pre_5=Light_5         # May (Pre-harvesting month)
            L_harv_5=Light_5+Light_6     # Jun = May + Jun
            
            L_Pre_6=Light_6         # Jun (Pre-harvesting month)
            L_harv_6=L_Pre_6+Light_7     # Jul = Jun + Jul
            
            L_Pre_7=Light_7         # Jul (Pre-harvesting month)
            L_harv_7=L_Pre_7+Light_8     # Aug = Jul + Aug
            
            L_Pre_8=Light_8         # Aug (Pre-harvesting month)
            L_harv_8=L_Pre_8+Light_9     # Sep = Aug + Sep
            
            L_Pre_9=Light_9         # Sep (Pre-harvesting month)
            L_harv_9=L_Pre_9+Light_10    # Oct = Sep + Oct
            
            L_Pre_10=Light_10       # oct (Pre-harvesting month)
            L_harv_10=L_Pre_10+Light_11  # Nov = Oct + Nov
            
            L_Pre_11=Light_11       # Nov (Pre-harvesting month)
            L_harv_11=L_Pre_11+Light_12  # Dec = Nov + Dec
            #L_Pre_12=Light_12
            #L_harv_12=L_Pre_12+Light_1   # Jan.
            
            CP_light_lettuce=[L_harv_1, L_harv_2, L_harv_3, L_harv_4, L_harv_5, L_harv_6, 
                            L_harv_7, L_harv_8, L_harv_9, L_harv_10, L_harv_11]
            
            LIGHT = []
            for sublist in CP_light_lettuce[:]:
                LIGHT.append(sublist[0])
            CP_LIGHT_LETTUCE= [round(number, 2) for number in LIGHT]

            # Writing to the desired cells in Excel
            if Country =="NL":
                sheetname = "NL_Lettuce"
            elif Country =="BE":
                sheetname = "BE_Lettuce"
            elif Country =="FR":
                sheetname = "FR_Lettuce"
            elif Country =="CH":     
                sheetname = "CH_Lettuce"
                
            cell_ranges = ["B6","B8","B10","B12", "B14", "B16", "B18", "B20", "B22", "B24", "B26"]
            Cultivations = ["Cultivation period1", "Cultivation period2", "Cultivation period3", "Cultivation period4", "Cultivation period5",
                            "Cultivation period6", "Cultivation period7", "Cultivation period8", "Cultivation period9", "Cultivation period10",
                            "Cultivation period11"]
            write_data_to_excel(filename, sheetname, cell_ranges, Cultivations)
            
            cell_ranges = ["C6","C7","C8","C9","C10", "C11", "C12", "C13", "C14", "C15", "C16", "C17", "C18", "C19", "C20", "C21", "C22", "C23",
                        "C24", "C25", "C26", "C27"]
            Cultivations = ["Heat(MJ/ha)","Lighting(kWh/ha)","Heat(MJ/ha)","Lighting(kWh/ha)","Heat(MJ/ha)","Lighting(kWh/ha)","Heat(MJ/ha)","Lighting(kWh/ha)",
                            "Heat(MJ/ha)","Lighting(kWh/ha)", "Heat(MJ/ha)","Lighting(kWh/ha)", "Heat(MJ/ha)","Lighting(kWh/ha)", "Heat(MJ/ha)","Lighting(kWh/ha)",
                            "Heat(MJ/ha)","Lighting(kWh/ha)", "Heat(MJ/ha)","Lighting(kWh/ha)", "Heat(MJ/ha)","Lighting(kWh/ha)"]
            write_data_to_excel(filename, sheetname, cell_ranges, Cultivations)         
                
            cell_ranges = ["B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5", "N5", "O5"]
            Month = ["Cultivation periods","Energy demand", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            write_data_to_excel(filename, sheetname, cell_ranges, Month) 
            
            cell_ranges = ["E6", "F8", "G10", "H12", "I14", "J16", "K18", "L20", "M22", "N24", "O26"]
            CP_Heat_Lettuce= CP_HEAT_LETTUCE  #[H1_harv_1, H1_harv_2, H1_harv_3, H1_harv_4, H1_harv_5, H1_harv_6, H1_harv_7, H1_harv_8, H1_harv_9]
            write_data_to_excel(filename, sheetname, cell_ranges,CP_Heat_Lettuce)
            cell_ranges = ["E7", "F9", "G11", "H13", "I15", "J17", "K19", "L21", "M23", "N25", "O27"]
            CP_Light_Lettuce= CP_LIGHT_LETTUCE  #[H1_harv_1, H1_harv_2, H1_harv_3, H1_harv_4, H1_harv_5, H1_harv_6, H1_harv_7, H1_harv_8, H1_harv_9]
            write_data_to_excel(filename, sheetname, cell_ranges,CP_Light_Lettuce)   
                
            print('done', i+'_'+j)



        ###########################################################
        ###########################################################
        ###########################################################
        
        # Cucumber (Cultivation period = 5 month and harvesting period =3 month)
        
        if Crop=="Cucumber" and Country in ["NL", "CH"]:
            
            # Heating demand pre-harvesting and harvesting month
            H1_Pre_1=Heat_1                              # Jan (Pre-harvesting month)
            H1_Pre_2=Heat_2                              # Feb (Pre-harvesting month)
            H1_harv_1=(H1_Pre_1+H1_Pre_2/2+Heat_3/3)     # Mar = Jan + 1/2 Feb + 1/3 Mar
            H1_harv_2=(H1_Pre_2/2+Heat_3/3+Heat_4/2)     # Apr = 1/2 Feb + 1/3 Mar + 1/2 Apr
            H1_harv_3=(Heat_3/3+Heat_4/2+Heat_5)         # May = 1/3 Mar + 1/2 Apr + May

            H2_Pre_1=Heat_2                              # Feb (Pre-harvesting month)
            H2_Pre_2=Heat_3                              # Mar (Pre-harvesting month)
            H2_harv_1=(H2_Pre_1+H2_Pre_2/2+Heat_4/3)     # Apr = Feb + 1/2 Mar + 1/3 Apr
            H2_harv_2=(H2_Pre_2/2+Heat_4/3+Heat_5/2)     # May = 1/2 Mar + 1/3 Apr + 1/2 May
            H2_harv_3=(Heat_4/3+Heat_5/2+Heat_6)         # Jun = 1/3 Apr + 1/2 May + Jun

            H3_Pre_1=Heat_3                              # Mar (Pre-harvesting month)
            H3_Pre_2=Heat_4                              # Apr (Pre-harvesting month)
            H3_harv_1=(H3_Pre_1+H3_Pre_2/2+Heat_5/3)     # May = Mar + 1/2 Apr + 1/3 May
            H3_harv_2=(H3_Pre_2/2+Heat_5/3+Heat_6/2)     # Jun = 1/2 Apr + 1/3 May + 1/2 Jun
            H3_harv_3=(Heat_5/3+Heat_6/2+Heat_7)         # Jul = 1/3 May + 1/2 Jun + Jul

            H4_Pre_1=Heat_4                              # Apr (Pre-harvesting month)
            H4_Pre_2=Heat_5                              # May (Pre-harvesting month)
            H4_harv_1=(H4_Pre_1+H4_Pre_2/2+Heat_6/3)     # Jun = Apr + 1/2 May + 1/3 Jun
            H4_harv_2=(H4_Pre_2/2+Heat_6/3+Heat_7/2)     # Jul = 1/2 May + 1/3 Jun + 1/2 Jul
            H4_harv_3=(Heat_6/3+Heat_7/2+Heat_8)         # Aug = 1/3 Jun + 1/2 Jul + Aug

            H5_Pre_1=Heat_5                              # May (Pre-harvesting month)
            H5_Pre_2=Heat_6                              # Jun (Pre-harvesting month)
            H5_harv_1=(H5_Pre_1+H5_Pre_2/2+Heat_7/3)     # Jul = May + 1/2 Jun + 1/3 Jul
            H5_harv_2=(H5_Pre_2/2+Heat_7/3+Heat_8/2)     # Aug = 1/2 Jun + 1/3 Jul + 1/2 Aug
            H5_harv_3=(Heat_7/3+Heat_8/2+Heat_9)         # Sep = 1/3 Jul + 1/2 Aug + Sep

            H6_Pre_1=Heat_6                              # Jun (Pre-harvesting month)
            H6_Pre_2=Heat_7                              # Jul (Pre-harvesting month)
            H6_harv_1=(H6_Pre_1+H6_Pre_2/2+Heat_8/3)     # Aug = Jun + 1/2 Jul + 1/3 Aug
            H6_harv_2=(H6_Pre_2/2+Heat_8/3+Heat_9/2)     # Sep = 1/2 Jul + 1/3 Aug + 1/2 Sep
            H6_harv_3=(Heat_8/3+Heat_9/2+Heat_10)        # Oct = 1/3 Aug + 1/2 Sep + Oct

            H7_Pre_1=Heat_7                              # Jul (Pre-harvesting month)
            H7_Pre_2=Heat_8                              # Aug (Pre-harvesting month)
            H7_harv_1=(H7_Pre_1+H7_Pre_2/2+Heat_9/3)     # Sep = Jul + 1/2 Aug + 1/3 Sep
            H7_harv_2=(H7_Pre_2/2+Heat_9/3+Heat_10/2)    # Oct = 1/2 Aug + 1/3 Sep + 1/2 Oct
            H7_harv_3=(Heat_9/3+Heat_10/2+Heat_11)       # Nov = 1/3 Sep + 1/2 Oct + Nov

            H8_Pre_1=Heat_8                              # Aug (Pre-harvesting month)
            H8_Pre_2=Heat_9                              # Sep (Pre-harvesting month)
            H8_harv_1=(H8_Pre_1+H8_Pre_2/2+Heat_10/3)    # Oct = Aug + 1/2 Sep + 1/3 Oct
            H8_harv_2=(H8_Pre_2/2+Heat_10/3+Heat_11/2)   # Nov = 1/2 Sep + 1/3 Oct + 1/2 Nov
            H8_harv_3=(Heat_10/3+Heat_11/2+Heat_12)      # Dec = 1/3 Oct + 1/2 Nov + Dec
        
            CP_heat_cucumber=[ 
                            H1_harv_1, H1_harv_2, H1_harv_3,
                            H2_harv_1, H2_harv_2, H2_harv_3,
                            H3_harv_1, H3_harv_2, H3_harv_3,
                            H4_harv_1, H4_harv_2, H4_harv_3,
                            H5_harv_1, H5_harv_2, H5_harv_3,
                            H6_harv_1, H6_harv_2, H6_harv_3,
                            H7_harv_1, H7_harv_2, H7_harv_3,
                            H8_harv_1, H8_harv_2, H8_harv_3,
                            ]
            
            HEAT = []
            for sublist in CP_heat_cucumber[:]:
                HEAT.append(sublist[0])
            CP_HEAT_CUCUMBER = [round(number, 2) for number in HEAT]
            
            
            # lighting demand pre-harvesting and harvesting month
            L1_Pre_1=Light_1                              # Jan (Pre-harvesting month)
            L1_Pre_2=Light_2                              # Feb (Pre-harvesting month)
            L1_harv_1=(L1_Pre_1+L1_Pre_2/2+Light_3/3)     # Mar = Jan + 1/2 Feb + 1/3 Mar
            L1_harv_2=(L1_Pre_2/2+Light_3/3+Light_4/2)    # Apr = 1/2 Feb + 1/3 Mar + 1/2 Apr
            L1_harv_3=(Light_3/3+Light_4/2+Light_5)       # May = 1/3 Mar + 1/2 Apr + May

            L2_Pre_1=Light_2                              # Feb (Pre-harvesting month)
            L2_Pre_2=Light_3                              # Mar (Pre-harvesting month)
            L2_harv_1=(L2_Pre_1+L2_Pre_2/2+Light_4/3)     # Apr = Feb + 1/2 Mar + 1/3 Apr
            L2_harv_2=(L2_Pre_2/2+Light_4/3+Light_5/2)    # May = 1/2 Mar + 1/3 Apr + 1/2 May
            L2_harv_3=(Light_4/3+Light_5/2+Light_6)       # Jun = 1/3 Apr + 1/2 May + Jun

            L3_Pre_1=Light_3                              # Mar (Pre-harvesting month)
            L3_Pre_2=Light_4                              # Apr (Pre-harvesting month)
            L3_harv_1=(L3_Pre_1+L3_Pre_2/2+Light_5/3)     # May = Mar + 1/2 Apr + 1/3 May
            L3_harv_2=(L3_Pre_2/2+Light_5/3+Light_6/2)    # Jun = 1/2 Apr + 1/3 May + 1/2 Jun
            L3_harv_3=(Light_5/3+Light_6/2+Light_7)        # Jul = 1/3 May + 1/2 Jun + Jul

            L4_Pre_1=Light_4                              # Apr (Pre-harvesting month)
            L4_Pre_2=Light_5                              # May (Pre-harvesting month)
            L4_harv_1=(L4_Pre_1+L4_Pre_2/2+Light_6/3)     # Jun = Apr + 1/2 May + 1/3 Jun
            L4_harv_2=(L4_Pre_2/2+Light_6/3+Light_7/2)    # Jul = 1/2 May + 1/3 Jun + 1/2 Jul
            L4_harv_3=(Light_6/3+Light_7/2+Light_8)       # Aug = 1/3 Jun + 1/2 Jul + Aug

            L5_Pre_1=Light_5                              # May (Pre-harvesting month)
            L5_Pre_2=Light_6                              # Jun (Pre-harvesting month)
            L5_harv_1=(L5_Pre_1+L5_Pre_2/2+Light_7/3)     # Jul = May + 1/2 Jun + 1/3 Jul
            L5_harv_2=(L5_Pre_2/2+Light_7/3+Light_8/2)    # Aug = 1/2 Jun + 1/3 Jul + 1/2 Aug
            L5_harv_3=(Light_7/3+Light_8/2+Light_9)       # Sep = 1/3 Jul + 1/2 Aug + Sep

            L6_Pre_1=Light_6                              # Jun (Pre-harvesting month)
            L6_Pre_2=Light_7                              # Jul (Pre-harvesting month)
            L6_harv_1=(L6_Pre_1+L6_Pre_2/2+Light_8/3)     # Aug = Jun + 1/2 Jul + 1/3 Aug
            L6_harv_2=(L6_Pre_2/2+Light_8/3+Light_9/2)    # Sep = 1/2 Jul + 1/3 Aug + 1/2 Sep
            L6_harv_3=(Light_8/3+Light_9/2+Light_10)      # Oct = 1/3 Aug + 1/2 Sep + Oct

            L7_Pre_1=Light_7                              # Jul (Pre-harvesting month)
            L7_Pre_2=Light_8                              # Aug (Pre-harvesting month)
            L7_harv_1=(L7_Pre_1+L7_Pre_2/2+Light_9/3)     # Sep = Jul + 1/2 Aug + 1/3 Sep
            L7_harv_2=(L7_Pre_2/2+Light_9/3+Light_10/2)   # Oct = 1/2 Aug + 1/3 Sep + 1/2 Oct
            L7_harv_3=(Light_9/3+Light_10/2+Light_11)     # Nov = 1/3 Sep + 1/2 Oct + Nov

            L8_Pre_1=Light_8                              # Aug (Pre-harvesting month)
            L8_Pre_2=Light_9                              # Sep (Pre-harvesting month)
            L8_harv_1=(L8_Pre_1+L8_Pre_2/2+Light_10/3)    # Oct = Aug + 1/2 Sep + 1/3 Oct
            L8_harv_2=(L8_Pre_2/2+Light_10/3+Light_11/2)  # Nov = 1/2 Sep + 1/3 Oct + 1/2 Nov
            L8_harv_3=(Light_10/3+Light_11/2+Light_12)    # Dec = 1/3 Oct + 1/2 Nov + Dec



            CP_light_cucumber=[
                            L1_harv_1, L1_harv_2, L1_harv_3,
                            L2_harv_1, L2_harv_2, L2_harv_3,
                            L3_harv_1, L3_harv_2, L3_harv_3,
                            L4_harv_1, L4_harv_2, L4_harv_3,
                            L5_harv_1, L5_harv_2, L5_harv_3,
                            L6_harv_1, L6_harv_2, L6_harv_3,
                            L7_harv_1, L7_harv_2, L7_harv_3,
                            L8_harv_1, L8_harv_2, L8_harv_3,
                            ]

            LIGHT = []
            for sublist in CP_light_cucumber[:]:
                LIGHT.append(sublist[0])
            CP_LIGHT_CUCUMBER= [round(number, 2) for number in LIGHT]


            # Writing to the desired cells in Excel
            if Country =="NL":
                sheetname = "NL_Cucumber"
            elif Country =="CH":     
                sheetname = "CH_Cucumber"
                
            cell_ranges = ["B6","B8","B10","B12","B14", "B16", "B18", "B20"]
            Cultivations = ["Cultivation period1", "Cultivation period2", "Cultivation period3", "Cultivation period4", "Cultivation period5",
                            "Cultivation period6", "Cultivation period7", "Cultivation period8"]
            write_data_to_excel(filename, sheetname, cell_ranges, Cultivations)
            
            cell_ranges = ["C6","C7","C8","C9","C10", "C11", "C12", "C13", "C14", "C15", "C16", "C17", "C18", "C19", "C20", "C21"]
            Energy_type = ["Heat(MJ/ha)","Lighting(kWh/ha)","Heat(MJ/ha)","Lighting(kWh/ha)","Heat(MJ/ha)","Lighting(kWh/ha)","Heat(MJ/ha)","Lighting(kWh/ha)",
                            "Heat(MJ/ha)","Lighting(kWh/ha)", "Heat(MJ/ha)","Lighting(kWh/ha)", "Heat(MJ/ha)","Lighting(kWh/ha)", "Heat(MJ/ha)","Lighting(kWh/ha)"]
            write_data_to_excel(filename, sheetname, cell_ranges, Energy_type)         
                
            cell_ranges = ["B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5", "N5", "O5"]
            Month = ["Cultivation periods","Energy demand", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            write_data_to_excel(filename, sheetname, cell_ranges, Month) 
            
            cell_ranges = ["F6", "G6", "H6", "G8", "H8", "I8", "H10", "I10", "J10", "I12", "J12", "K12", 
                        "J14", "K14", "L14", "K16", "L16", "M16", "L18", "M18", "N18", "M20", "N20", "O20"]
            CP_Heat_Cucumber= CP_HEAT_CUCUMBER  
            write_data_to_excel(filename, sheetname, cell_ranges, CP_Heat_Cucumber)
            
            cell_ranges = ["F7", "G7", "H7", "G9", "H9", "I9", "H11", "I11", "J11", "I13", "J13", "K13", 
                        "J15", "K15", "L15", "K17", "L17", "M17", "L19", "M19", "N19", "M21", "N21", "O21"]
            CP_Light_Cucumber= CP_LIGHT_CUCUMBER  
            write_data_to_excel(filename, sheetname, cell_ranges, CP_Light_Cucumber)   
                
            print('done', i+'_'+j)



        ###########################################################
        ###########################################################
        ###########################################################
        
        # Strawberry
        
        if Crop=="Strawberry" and Country in ["NL", "BE", "CH"]:
            
            # Heating demand pre-harvesting and harvesting month
            H1_Pre_1=Heat_1                                       # Jan (Pre-harvesting month)
            H1_Pre_2=Heat_2                                       # Feb (Pre-harvesting month)
            H1_Pre_3=Heat_3                                       # Mar (Pre-harvesting month)
            H1_harv_1=(H1_Pre_1+H1_Pre_2/2+H1_Pre_3/3+Heat_4/4)   # Apr = Jan + 1/2 Feb + 1/3 Mar + 1/4 Apr
            H1_harv_2=(H1_Pre_2/2+H1_Pre_3/3+Heat_4/4+Heat_5/4)   # May = 1/2 Feb + 1/3 Mar + 1/4 Apr + 1/4 May
            H1_harv_3=(H1_Pre_3/3+Heat_4/4+Heat_5/4+Heat_6/4)     # Jun = 1/3 Mar + 1/4 Apr + 1/4 May + 1/4 Jun
            H1_harv_4=(Heat_4/4+Heat_5/4+Heat_6/4+Heat_7/4)       # Jul = 1/4 Apr + 1/4 May + 1/4 Jun + 1/4 Jul
            H1_harv_5=(Heat_5/4+Heat_6/4+Heat_7/4+Heat_8/4)       # Aug = 1/4 May + 1/4 Jun + 1/4 Jul + 1/4 Aug
            H1_harv_6=(Heat_6/4+Heat_7/4+Heat_8/4+Heat_9/4)       # Sep = 1/4 Jun + 1/4 Jul + 1/4 Aug + 1/4 Sep
            H1_harv_7=(Heat_7/4+Heat_8/4+Heat_9/4+Heat_10/3)      # Oct = 1/4 Jul + 1/4 Aug + 1/4 Sep + 1/3 Oct
            H1_harv_8=(Heat_8/4+Heat_9/4+Heat_10/3+Heat_11/2)     # Nov = 1/4 Aug + 1/4 Sep + 1/3 Oct + 1/2 Nov
            H1_harv_9=(Heat_9/4+Heat_10/3+Heat_11/2+Heat_12)      # Dec = 1/4 Sep + 1/3 Oct + 1/2 Nov + Dec
            
            CP_heat_strawberry= [H1_harv_1, H1_harv_2, H1_harv_3, H1_harv_4, H1_harv_5,
                                H1_harv_6, H1_harv_7, H1_harv_8, H1_harv_9]

            HEAT = []
            for sublist in CP_heat_strawberry[:]:
                HEAT.append(sublist[0])
            CP_HEAT_STRAWBERRY = [round(number, 2) for number in HEAT]


            # Lighting demand pre-harvesting and harvesting month
            L1_Pre_1=Light_1                                          # Jan (Pre-harvesting month)
            L1_Pre_2=Light_2                                          # Feb (Pre-harvesting month)
            L1_Pre_3=Light_3                                          # Mar (Pre-harvesting month)
            L1_harv_1=(L1_Pre_1+L1_Pre_2/2+L1_Pre_3/3+Light_4/4)      # Apr = Jan + 1/2 Feb + 1/3 Mar + 1/4 Apr
            L1_harv_2=(L1_Pre_2/2+L1_Pre_3/3+Light_4/4+Light_5/4)     # May = 1/2 Feb + 1/3 Mar + 1/4 Apr + 1/4 May
            L1_harv_3=(L1_Pre_3/3+Light_4/4+Light_5/4+Light_6/4)      # Jun = 1/3 Mar + 1/4 Apr + 1/4 May + 1/4 Jun
            L1_harv_4=(Light_4/4+Light_5/4+Light_6/4+Light_7/4)       # Jul = 1/4 Apr + 1/4 May + 1/4 Jun + 1/4 Jul
            L1_harv_5=(Light_5/4+Light_6/4+Light_7/4+Light_8/4)       # Aug = 1/4 May + 1/4 Jun + 1/4 Jul + 1/4 Aug
            L1_harv_6=(Light_6/4+Light_7/4+Light_8/4+Light_9/4)       # Sep = 1/4 Jun + 1/4 Jul + 1/4 Aug + 1/4 Sep
            L1_harv_7=(Light_7/4+Light_8/4+Light_9/4+Light_10/3)      # Oct = 1/4 Jul + 1/4 Aug + 1/4 Sep + 1/3 Oct
            L1_harv_8=(Light_8/4+Light_9/4+Light_10/3+Light_11/2)     # Nov = 1/4 Aug + 1/4 Sep + 1/3 Oct + 1/2 Nov
            L1_harv_9=(Light_9/4+Light_10/3+Light_11/2+Light_12)      # Dec = 1/4 Sep + 1/3 Oct + 1/2 Nov + Dec

            CP_light_strawberry= [L1_harv_1, L1_harv_2, L1_harv_3, L1_harv_4, L1_harv_5, 
                                L1_harv_6, L1_harv_7, L1_harv_8, L1_harv_9]

            LIGHT = []
            for sublist in CP_light_strawberry[:]:
                LIGHT.append(sublist[0])
            CP_LIGHT_STRAWBERRY= [round(number, 2) for number in LIGHT]


            # Writing to the desired cells in Excel
            if Country =="NL":
                sheetname = "NL_Strawberry"
            elif Country =="BE":     
                sheetname = "BE_Strawberry"
            elif Country =="CH":     
                sheetname = "CH_Strawberry"
            cell_ranges = ["B5","C5","C6","C7","D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5", "N5", "O5"]
            Month = ["Cultivation period","Energy demand","Heat(MJ/ha)","Lighting(kWh/ha)","Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            write_data_to_excel(filename, sheetname, cell_ranges, Month) 
            cell_ranges = ["G6", "H6", "I6", "J6", "K6", "L6", "M6", "N6", "O6"]
            CP_Heat_Strawberry= CP_HEAT_STRAWBERRY  
            write_data_to_excel(filename, sheetname, cell_ranges, CP_Heat_Strawberry )
            cell_ranges = ["G7", "H7", "I7", "J7", "K7", "L7", "M7", "N7", "O7"]
            CP_Light_Strawberry= CP_LIGHT_STRAWBERRY  
            write_data_to_excel(filename, sheetname, cell_ranges,CP_Light_Strawberry) 
            print('done', i+'_'+j)


print("All Done")